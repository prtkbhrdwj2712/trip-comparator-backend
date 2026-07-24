"""
Parses the dispatch-summary-style Excel file (same layout as the files used
to build the first version of the dashboard) into plain dicts ready to be
upserted into TripBaseline/StopBaseline or TripConfirmed/StopConfirmed.

If the underlying export format ever changes, this is the one place to
update - everything downstream (diff engine, API, dashboard) is unaffected
as long as this keeps returning the same shape.
"""
import openpyxl
from datetime import datetime


def _stringify(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return v


TRIP_FIELD_MAP = {
    "plan id": "plan_id",
    "trip id": "trip_id",
    "trip name": "trip_name",
    "trip date": "trip_date",
    "vehicle category": "vehicle_category",
    "vehicle id": "vehicle_id",
    "driver name": "driver_name",
    "planned trip distance(km)": "trip_distance_km",
    "planned trip duration(h)": "trip_duration_h",
    "trip weight(kg)": "trip_weight_kg",
    "trip volume(cm3)": "trip_volume_cm3",
    "weight utilization": "weight_utilization",
    "space utilization": "space_utilization",
    "distance utilization": "distance_utilization",
    "time utilization": "time_utilization",
    "Trip Cost": "trip_cost",
    "status": "status",
}


def parse_dispatch_workbook(file_like):
    """
    file_like: a file path or an in-memory binary stream (e.g. from an
    uploaded multipart file).
    Returns: dict[trip_id] -> {..trip fields.., "stops": [ {..stop..}, ... ]}
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}

    required = ["trip id", "trip activity", "ship to (code)"]
    for r in required:
        if r not in idx:
            raise ValueError(f"Expected column '{r}' not found in workbook headers: {headers}")

    trips = {}
    for row_num in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
        tid = row[idx["trip id"]]
        if tid is None:
            continue

        if tid not in trips:
            t = {}
            for src, dest in TRIP_FIELD_MAP.items():
                if src in idx:
                    t[dest] = _stringify(row[idx[src]])
            t["stops"] = {}
            trips[tid] = t

        real_stop_code = row[idx["ship to (code)"]]
        activity = row[idx["trip activity"]]

        if activity == "Drop" and not real_stop_code:
            # A blank ship-to code on a Drop row is the vehicle's
            # return-to-depot leg at the end of the route (its "dealer name"
            # is always blank too) - not a real delivery stop. Previously
            # this fell back to using the raw depot address as a fake dealer
            # code, which inflated the dealer count and polluted the route
            # view with a garbage "dealer". Skip it entirely.
            continue

        stop_code = real_stop_code
        key = f"{activity}_{stop_code}"
        stops = trips[tid]["stops"]
        if key not in stops:
            stops[key] = {
                "activity": activity,
                "ship_to_code": stop_code,
                "ship_to_name": _stringify(row[idx["ship to (name)"]]) if "ship to (name)" in idx else None,
                "sequence": _stringify(row[idx["sequence"]]) if "sequence" in idx else None,
                "arrival": _stringify(row[idx["planned arrival"]]) if "planned arrival" in idx else None,
                "reference_order_number": _stringify(row[idx["reference order number"]]) if "reference order number" in idx else None,
                "weight_kg": 0.0,
            }
        # weight(kg) is a per-SKU-line value - a stop has one row per SKU
        # line, so the true weight delivered at this stop is the SUM across
        # all its lines, not just whichever row happened to be seen first
        # (that was a latent bug - only the first line's weight was kept).
        if "weight(kg)" in idx:
            line_weight = row[idx["weight(kg)"]]
            if isinstance(line_weight, (int, float)):
                stops[key]["weight_kg"] += line_weight

    for tid, t in trips.items():
        t["no_of_stops"] = len([k for k in t["stops"] if k.startswith("Drop")])
        t["stops"] = list(t["stops"].values())

    return trips
